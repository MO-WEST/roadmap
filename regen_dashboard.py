#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Aktualisiert die eingebetteten Daten in roadmap-dashboard.html aus
Entwicklungs_Roadmap_2026-2028.xlsx.

Aufruf:  python regen_dashboard.py
Voraussetzung:  pip install openpyxl
"""
import io, os, re, json, warnings
import openpyxl

warnings.filterwarnings("ignore")
HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Entwicklungs_Roadmap_2026-2028.xlsx")
HTML = os.path.join(HERE, "roadmap-dashboard.html")
YEARS = [2026, 2027, 2028]
BAD = ("∑", "ℹ", "🟢", "🟡", "MoSCoW", "Value:", "Größen")


def num(v):
    return v if isinstance(v, (int, float)) else 0


def extract():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    data = {}
    for y in YEARS:
        dash = wb[f"Dashboard{y}"]
        road = wb[f"Roadmap{y}"]
        drow = lambda r: [dash.cell(r, c).value for c in range(1, 7)]
        crow = lambda r: [num(dash.cell(r, c).value) for c in range(2, 6)]   # Q1..Q4
        mrow = lambda r: [num(dash.cell(r, c).value) for c in range(2, 7)]   # Q1..Q4 + Gesamt
        rec = {
            "value": {"Strengthen": drow(5)[1:6], "Innovation": drow(7)[1:6], "Repair": drow(9)[1:6]},
            "totalProjects": drow(11)[1:6],
            "capTotal": crow(16), "capDesktop": crow(17), "capWeb": crow(18),
            "planTotal": crow(19), "effDesktop": crow(24), "effWeb": crow(25),
            "moscow": {"Must": mrow(32), "Should": mrow(33), "Could": mrow(34), "Won't": mrow(35)},
            "team": {"WEB": mrow(39), "Desktop": mrow(40), "W->D": mrow(41), "D->W": mrow(42)},
            "poBacklog": mrow(46),
        }
        # Projektlisten pro Quartal aus dem Roadmap-Blatt (Header Zeile 17, Daten ab 18)
        qcols = {"Q1": 2, "Q2": 6, "Q3": 10, "Q4": 14}
        projects = {q: [] for q in qcols}
        sumrow = road.max_row
        for r in range(18, road.max_row + 1):
            b = road.cell(r, 2).value
            if isinstance(b, str) and b.strip().startswith("∑"):
                sumrow = r
                break
        for q, c0 in qcols.items():
            for r in range(18, sumrow):
                t = road.cell(r, c0).value
                if isinstance(t, str) and t.strip() and not any(t.strip().startswith(x) for x in BAD):
                    s = road.cell(r, c0 + 1).value
                    v = road.cell(r, c0 + 2).value
                    projects[q].append({
                        "t": t.strip(),
                        "s": (str(s) if s not in (None, 0) else ""),
                        "v": (str(v) if v not in (None, 0) else ""),
                    })
        rec["projects"] = projects
        data[y] = rec
    return data


def main():
    data = extract()
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    html = io.open(HTML, encoding="utf-8").read()
    new = re.sub(r"const DATA = .*?;\s*$",
                 "const DATA = " + payload + ";",
                 html, count=1, flags=re.MULTILINE)
    if new == html:
        raise SystemExit("Marker 'const DATA = …;' nicht gefunden – HTML unverändert.")
    io.open(HTML, "w", encoding="utf-8", newline="\n").write(new)
    print("OK – roadmap-dashboard.html aktualisiert.")
    for y in YEARS:
        print(f"  {y}: " + ", ".join(f"{q}={len(data[y]['projects'][q])}" for q in ("Q1", "Q2", "Q3", "Q4")))


if __name__ == "__main__":
    main()
